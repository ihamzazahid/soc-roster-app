from io import BytesIO
from functools import wraps
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from flask import redirect, url_for, flash
from flask_login import current_user

# --------------------------------------------------------------------------
# ROLE-BASED ACCESS CONTROL (RBAC) DECORATORS
# --------------------------------------------------------------------------

def role_required(*roles):
    """
    Decorator enforcing that current_user has one of the specified system roles.
    Supports string role fields (user.role = 'Admin') or Relationship objects (user.role.name = 'Admin').
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                flash('Please log in to access this page.', 'warning')
                return redirect(url_for('main.login'))

            user_role = ''
            if hasattr(current_user, 'role') and current_user.role:
                user_role = current_user.role.name if hasattr(current_user.role, 'name') else str(current_user.role)

            if user_role not in roles:
                flash('Unauthorized access privileges. Admin access required.', 'danger')
                return redirect(url_for('main.dashboard'))

            return f(*args, **kwargs)
        return decorated_function
    return decorator


def admin_required(f):
    """Shorthand decorator specifically for Admin-only routes."""
    return role_required('Admin', 'SOC_Manager')(f)

# --------------------------------------------------------------------------
# EXCEL ROSTER EXPORTER
# --------------------------------------------------------------------------

def export_roster_to_excel(roster_data, year, month):
    """Generates a formatted Excel spreadsheet for a month's roster."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Roster {month:02d}-{year}"

    header_fill = PatternFill(start_color="1A252F", end_color="1A252F", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    headers = ["Date", "Analyst", "Role / Tier", "Shift Assigned"]
    ws.append(headers)

    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for idx, item in enumerate(roster_data, start=2):
        ws.cell(row=idx, column=1, value=item['Date']).alignment = center
        ws.cell(row=idx, column=2, value=item['Analyst'])
        ws.cell(row=idx, column=3, value=item['Role']).alignment = center
        
        shift_cell = ws.cell(row=idx, column=4, value=item['Shift'])
        shift_cell.alignment = center
        
        for c in range(1, 5):
            ws.cell(row=idx, column=c).border = border

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream